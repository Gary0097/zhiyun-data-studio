# -*- coding: utf-8 -*-
"""CSV and XLSX parsing independent from the QwenPaw HTTP layer."""

from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any

MAX_ROWS = 10000


def _clean_rows(rows: list[list[Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    if not rows:
        raise ValueError("文件中没有数据")
    headers = [str(value or "").strip() for value in rows[0]]
    active_headers = [header for header in headers if header]
    if not active_headers:
        raise ValueError("第一行必须是字段名称")
    if len(active_headers) != len(set(active_headers)):
        raise ValueError("字段名称不能重复")
    records: list[dict[str, Any]] = []
    for values in rows[1 : MAX_ROWS + 1]:
        if not any(value not in (None, "") for value in values):
            continue
        records.append(
            {
                header: values[index] if index < len(values) else None
                for index, header in enumerate(headers)
                if header
            }
        )
    return active_headers, records


def parse_table(filename: str, content: bytes) -> dict[str, Any]:
    """Parse CSV or XLSX bytes into standard rows for Data Core preview."""
    suffix = Path(filename).suffix.casefold()
    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        matrix = list(csv.reader(io.StringIO(text)))
        headers, rows = _clean_rows(matrix)
        return {
            "filename": filename,
            "sheet": None,
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
        }
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
        headers, rows = _clean_rows(matrix)
        return {
            "filename": filename,
            "sheet": sheet.title,
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
        }
    raise ValueError("仅支持 .xlsx 和 .csv 文件")
